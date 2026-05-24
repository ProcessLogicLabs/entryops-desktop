"""File-format adapters that produce a normalized flat-text representation
of a supplier's ISF 10+2 data so the same regex-based template can parse
either a PDF or an Excel container.

Two known supplier formats:
  - PDF: standardized "Importer Security Filing (ISF) Information Sheet" form
    with all 17 fields laid out in labeled rows. Read via pdfplumber.
  - XLS/XLSX: a 2-column "ISF Template" worksheet (label in column A, value
    in column B) carrying the same 17 fields. Excel stores dates as serial
    numbers and ZIPs/HTS codes as floats; we normalize on the way out.

The output is a single string formatted to look like the PDF flat-text so
`Entryops.templates.isf_10_plus_2.ISF10Plus2Template` can consume it
unchanged.
"""

from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path
from typing import Any, List, Tuple


def _coerce_value(v: Any, label: str = "") -> str:
    """Convert an Excel cell value to a clean string.

    - Excel serial dates (numeric > 25000) → MM/DD/YYYY when the label hints
      a date field (ETD/ETA/sailing/arrival).
    - Trailing '.0' on integer-floats stripped (ZIP codes, HTS, port codes).
    - Strings stripped of surrounding whitespace.
    """
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, (int, float)):
        is_date_label = bool(re.search(r"\b(ETD|ETA|sailing|arrival|date)\b", label, re.I))
        if is_date_label and 25000 < float(v) < 100000:
            try:
                d = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=float(v))
                return d.strftime("%m/%d/%Y")
            except (OverflowError, ValueError):
                pass
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return str(v)
    return str(v).strip()


def _read_pdf_text(path: Path, db_path: Path | None = None) -> str:
    """Read a PDF as a single concatenated text blob via pdfplumber.

    If pdfplumber returns no usable text (image-only / scanned PDF), fall back
    to the shared EntryOps OCR engine (Tesseract by default). The OCR config
    is read from the same `billing_settings` table the rest of the app uses
    so admin overrides apply uniformly.
    """
    import pdfplumber
    parts: List[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t:
                parts.append(t)
    text = "\n".join(parts)
    if text.strip():
        return text
    return _ocr_pdf_text(path, db_path)


def _ocr_pdf_text(path: Path, db_path: Path | None) -> str:
    """Run the configured EntryOps OCR backend; return '' on any failure."""
    try:
        from entryops.ocrmill_processor import _OCRConfigAdapter
        from entryops.ocr import get_ocr_engine
    except ImportError:
        try:
            from ocrmill_processor import _OCRConfigAdapter
            from ocr import get_ocr_engine
        except ImportError:
            return ""

    config = _OCRConfigAdapter(db_path)
    if not config.get("ocr.enabled", False):
        return ""

    try:
        engine = get_ocr_engine(config)
        dpi = int(config.get("ocr.tesseract.dpi", 300) or 300)
        preprocess = config.get("ocr.tesseract.preprocess", None)
        return engine.ocr_pdf(path, dpi=dpi, preprocess=preprocess) or ""
    except (NotImplementedError, OSError, RuntimeError):
        return ""
    except Exception:
        return ""


def _read_xls_rows(path: Path) -> List[Tuple[str, str]]:
    """Return (label, value) pairs from a 2-column ISF template (.xls or .xlsx).

    Picks the sheet most likely to be the ISF data: prefer one whose name
    contains "ISF", else the first sheet that has 'Importer Security Filing'
    in any cell of column A.
    """
    suffix = path.suffix.lower()
    rows: List[Tuple[Any, Any]] = []

    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path), formatting_info=False)
        sheet = None
        for s in wb.sheets():
            if "ISF" in s.name.upper() and s.nrows > 0:
                sheet = s
                break
        if sheet is None:
            for s in wb.sheets():
                if s.nrows == 0:
                    continue
                for r in range(min(s.nrows, 5)):
                    cell = s.cell_value(r, 0) if s.ncols > 0 else ""
                    if isinstance(cell, str) and "Importer Security Filing" in cell:
                        sheet = s
                        break
                if sheet:
                    break
        if sheet is None:
            return []
        for r in range(sheet.nrows):
            a = sheet.cell_value(r, 0) if sheet.ncols > 0 else ""
            b = sheet.cell_value(r, 1) if sheet.ncols > 1 else ""
            rows.append((a, b))
    elif suffix == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet = None
        for name in wb.sheetnames:
            if "ISF" in name.upper():
                sheet = wb[name]
                break
        if sheet is None:
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                    if row and isinstance(row[0], str) and "Importer Security Filing" in row[0]:
                        sheet = ws
                        break
                if sheet:
                    break
        if sheet is None:
            return []
        for row in sheet.iter_rows(values_only=True):
            a = row[0] if len(row) > 0 else ""
            b = row[1] if len(row) > 1 else ""
            rows.append((a, b))
    else:
        return []

    pairs: List[Tuple[str, str]] = []
    for a, b in rows:
        a_str = a.strip() if isinstance(a, str) else ("" if a is None else str(a).strip())
        b_str = _coerce_value(b, label=a_str)
        if a_str or b_str:
            pairs.append((a_str, b_str))
    return pairs


def _xls_pairs_to_isf_text(pairs: List[Tuple[str, str]]) -> str:
    """Reformat 2-column ISF template rows into the flat-text shape that
    matches what pdfplumber emits for the PDF version of the form, so the
    same regex template can parse either source.

    Output rules:
      - Lines whose label starts with "1." through "17." emit "<label> <value>".
      - Address-block labels (NAME OF COMPANY:, ADDRESS:, CITY:, etc.) emit
        "<label> <value>" as their own line.
      - Other rows pass through (helps `can_process` see the title sentinel).
    """
    out: List[str] = []
    NUMBERED = re.compile(r"^\s*(\d+)\s*\.\s*(.+)$")
    ADDR_LABEL = re.compile(r"^\s*(NAME OF COMPANY|ADDRESS|CITY|STATE\s*/\s*PROVINCE\s*/\s*ZIP CODE|COUNTRY)\s*:\s*$", re.I)
    for label, value in pairs:
        m = NUMBERED.match(label)
        if m:
            normalized_label = re.sub(r"\s+", " ", label.strip())
            if value:
                out.append(f"{normalized_label} {value}")
            else:
                out.append(normalized_label)
            continue
        if ADDR_LABEL.match(label):
            normalized_label = re.sub(r"\s+", " ", label.strip().rstrip(":")) + ":"
            if value:
                out.append(f"{normalized_label} {value}")
            else:
                out.append(normalized_label)
            continue
        if label:
            out.append(label)
        elif value:
            out.append(value)
    return "\n".join(out)


def load_isf_text(path: str | Path, db_path: str | Path | None = None) -> str:
    """Load a supplier ISF source file (PDF or Excel) and return its
    normalized flat-text form for template-based extraction.

    db_path is used for the OCR fallback config lookup; if omitted, OCR
    runs with hardcoded defaults.

    Line endings are normalized to '\\n' for cross-platform regex stability.

    Raises ValueError on unsupported extensions.
    """
    p = Path(path)
    suffix = p.suffix.lower()
    db = Path(db_path) if db_path else None
    if suffix == ".pdf":
        text = _read_pdf_text(p, db)
    elif suffix in (".xls", ".xlsx"):
        pairs = _read_xls_rows(p)
        text = _xls_pairs_to_isf_text(pairs)
    else:
        raise ValueError(f"Unsupported ISF source format: {suffix}")
    return text.replace("\r\n", "\n").replace("\r", "\n")

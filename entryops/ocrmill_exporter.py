"""
OCRMill Profile Exporter

Loads output profiles from the database and exports enriched DataFrames
to XLSX with material-type color coding, matching the main EntryOps
export formatting.
"""

import json
import logging
import shutil
import sqlite3
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font as ExcelFont, PatternFill

logger = logging.getLogger(__name__)

# Default column order (matches entryops.py line 13354)
DEFAULT_OUTPUT_COLUMN_ORDER = [
    'Product No', 'ValueUSD', 'HTSCode', 'MID', 'Qty1', 'Qty2',
    'GrossWeight',
    'DecTypeCd', 'CountryofMelt', 'CountryOfCast', 'PrimCountryOfSmelt',
    'DeclarationFlag', 'SteelPercentage', 'AluminumPercentage', 'CopperPercentage',
    'WoodPercentage', 'AutoPercentage', 'NonSteelPercentage', '232_Status',
    'Ch99Heading', 'Ch99Rate', 'MetalWeightPct', 'MetalWeightKG',
    'Sec122HTS',     # Section 122 (Reciprocal Tariffs) heading: 9903.03.06 if Sec 232 applies, else 9903.03.01
    'CustomerRef',
    'ReviewFlag',  # post-export operator follow-up: '' / INCOMPLETE / NOT_IN_DB / UNMAPPED_ALIAS
]


class ProfileExporter:
    """
    Loads output profiles from the database and exports enriched DataFrames
    to formatted XLSX files.
    """

    def __init__(self, db_path: Path, output_dir: Path, log_callback=None):
        self.db_path = Path(db_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log_callback = log_callback

    def _log(self, msg: str):
        logger.info(msg)
        if self.log_callback:
            self.log_callback(msg)

    def get_available_profiles(self) -> List[str]:
        """Get list of saved output profile names from the database."""
        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("SELECT profile_name FROM output_column_mappings ORDER BY profile_name")
            profiles = [row[0] for row in c.fetchall()]
            conn.close()
            return profiles
        except Exception as e:
            logger.error(f"Failed to load output profiles: {e}")
            return []

    def load_profile(self, profile_name: str) -> Optional[Dict]:
        """
        Load an output profile from the database.

        Returns dict with keys: column_mapping, column_order, column_visibility, split_by_invoice
        Returns None if profile not found.
        """
        if not profile_name:
            return None

        try:
            conn = sqlite3.connect(str(self.db_path))
            c = conn.cursor()
            c.execute("SELECT mapping_json FROM output_column_mappings WHERE profile_name=?", (profile_name,))
            row = c.fetchone()
            conn.close()

            if not row:
                return None

            profile_data = json.loads(row[0])

            # Handle new format (nested) vs old format (flat mapping)
            if 'column_mapping' in profile_data:
                return {
                    'column_mapping': profile_data.get('column_mapping', {}),
                    'column_order': profile_data.get('column_order', DEFAULT_OUTPUT_COLUMN_ORDER),
                    'column_visibility': profile_data.get('column_visibility', {}),
                    'split_by_invoice': profile_data.get('split_by_invoice', False),
                }
            else:
                return {
                    'column_mapping': profile_data,
                    'column_order': DEFAULT_OUTPUT_COLUMN_ORDER.copy(),
                    'column_visibility': {},
                    'split_by_invoice': False,
                }
        except Exception as e:
            logger.error(f"Failed to load profile '{profile_name}': {e}")
            return None

    def build_export_df(self, df_enriched: pd.DataFrame, profile: Dict) -> Tuple[pd.DataFrame, List[str]]:
        """
        Apply output profile to enriched DataFrame: column order, pm: columns,
        blank: columns, visibility, and renaming.

        Returns (df_out, cols) ready for XLSX writing.
        """
        df_out = df_enriched.copy()

        column_order = profile.get('column_order', DEFAULT_OUTPUT_COLUMN_ORDER)
        column_mapping = profile.get('column_mapping', {})
        column_visibility = profile.get('column_visibility', {})

        # Build material masks BEFORE converting ratios to percentage strings
        steel_mask = df_out['_232_flag'].fillna('').str.contains('232_Steel', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        aluminum_mask = df_out['_232_flag'].fillna('').str.contains('232_Aluminum', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        copper_mask = df_out['_232_flag'].fillna('').str.contains('232_Copper', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        wood_mask = df_out['_232_flag'].fillna('').str.contains('232_Wood', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        auto_mask = df_out['_232_flag'].fillna('').str.contains('232_Auto', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        non232_mask = df_out['_232_flag'].fillna('').str.contains('Non_232', case=False, na=False) if '_232_flag' in df_out.columns else pd.Series([False] * len(df_out))
        # Cast iron / zero-metal-232: rows whose Ch99Heading is 9903.82.01.
        # Distinct color so filers can identify them at a glance even though
        # the 232 Status still reads 232_Steel / 232_Aluminum.
        zero_metal_232_mask = df_out['Ch99Heading'].fillna('').astype(str).str.strip() == '9903.82.01' if 'Ch99Heading' in df_out.columns else pd.Series([False] * len(df_out))

        # Sec301 exclusion mask
        sec301_mask = pd.Series([False] * len(df_out))
        if '_sec301_exclusion' in df_out.columns:
            sec301_mask = df_out['_sec301_exclusion'].fillna('').astype(str).str.strip().ne('') & \
                          ~df_out['_sec301_exclusion'].fillna('').astype(str).str.contains('nan|None', case=False, na=False)

        # Dual declaration — only compute if in the profile
        if 'DualDeclaration' in column_order:
            if 'SteelPercentage' in df_out.columns and 'AluminumPercentage' in df_out.columns:
                dual_dec_mask = (pd.to_numeric(df_out['SteelPercentage'], errors='coerce').fillna(0) > 0) & \
                                (pd.to_numeric(df_out['AluminumPercentage'], errors='coerce').fillna(0) > 0)
            else:
                dual_dec_mask = pd.Series([False] * len(df_out))
            df_out['DualDeclaration'] = dual_dec_mask.apply(lambda x: '07 & 08' if x else '')
        else:
            dual_dec_mask = pd.Series([False] * len(df_out))

        # Convert ratio values to percentage strings
        for ratio_col in ['SteelPercentage', 'AluminumPercentage', 'CopperPercentage', 'WoodPercentage', 'AutoPercentage', 'NonSteelPercentage']:
            if ratio_col in df_out.columns:
                df_out[ratio_col] = pd.to_numeric(df_out[ratio_col], errors='coerce').fillna(0).round(1).astype(str) + "%"

        # Set 232_Status from _232_flag
        if '_232_flag' in df_out.columns:
            df_out['232_Status'] = df_out['_232_flag'].fillna('')

        # Add pm: (parts_master) columns
        pm_cols_needed = [name for name in column_order if name.startswith('pm:')]
        if pm_cols_needed and 'Product No' in df_out.columns:
            try:
                pm_fields = [name[3:] for name in pm_cols_needed]
                conn = sqlite3.connect(str(self.db_path))
                fields_sql = ', '.join(['part_number'] + pm_fields)
                pm_data = pd.read_sql(f"SELECT {fields_sql} FROM parts_master", conn)
                conn.close()
                pm_data['part_number'] = pm_data['part_number'].astype(str).str.strip().str.upper()
                pm_data = pm_data.drop_duplicates(subset='part_number', keep='first')
                pm_lookup = pm_data.set_index('part_number')
                product_nos = df_out['Product No'].astype(str).str.strip().str.upper()
                for pm_field in pm_fields:
                    col_key = f'pm:{pm_field}'
                    if pm_field in pm_lookup.columns:
                        df_out[col_key] = product_nos.map(pm_lookup[pm_field]).fillna('')
                    else:
                        df_out[col_key] = ''
            except Exception as e:
                logger.warning(f"Failed to fetch parts_master fields for export: {e}")
                for name in pm_cols_needed:
                    df_out[name] = ''

        # Add blank spacer columns
        for name in column_order:
            if name.startswith('blank:'):
                df_out[name] = ''

        all_columns = list(column_order)

        # Filter columns by visibility
        ratio_columns = ['SteelPercentage', 'AluminumPercentage', 'CopperPercentage', 'WoodPercentage', 'AutoPercentage', 'NonSteelPercentage']
        cols = []
        for col in all_columns:
            if col in ratio_columns:
                # Check visibility from profile or DB
                is_visible = column_visibility.get(col, True)
                if not is_visible:
                    # Also check DB
                    try:
                        conn = sqlite3.connect(str(self.db_path))
                        c = conn.cursor()
                        c.execute("SELECT value FROM app_config WHERE key = ?", (f'export_col_visible_{col}',))
                        row = c.fetchone()
                        conn.close()
                        if row:
                            is_visible = row[0] == 'True'
                    except:
                        pass
                if is_visible:
                    cols.append(col)
            else:
                cols.append(col)

        # Filter out columns that don't exist in DataFrame
        cols = [col for col in cols if col in df_out.columns]

        if not cols:
            cols = [c for c in DEFAULT_OUTPUT_COLUMN_ORDER if c in df_out.columns]

        # Apply custom column renaming
        if column_mapping:
            rename_dict = {}
            blank_counter = 0
            for col in cols:
                if col in column_mapping and column_mapping[col] != col:
                    display_name = column_mapping[col]
                    if col.startswith('blank:') and not display_name.strip():
                        blank_counter += 1
                        display_name = ' ' * blank_counter
                    rename_dict[col] = display_name

            if rename_dict:
                df_out = df_out.rename(columns=rename_dict)
                new_cols = []
                for col in cols:
                    if col in rename_dict:
                        new_cols.append(rename_dict[col])
                    elif col in column_mapping and column_mapping[col] != col:
                        new_cols.append(column_mapping[col])
                    else:
                        new_cols.append(col)
                cols = new_cols

        masks = {
            'steel': steel_mask,
            'aluminum': aluminum_mask,
            'copper': copper_mask,
            'wood': wood_mask,
            'auto': auto_mask,
            'non232': non232_mask,
            'sec301': sec301_mask,
            'dual_dec': dual_dec_mask,
            'zero_metal_232': zero_metal_232_mask,
        }

        return df_out, cols, masks

    def write_xlsx(self, df_out: pd.DataFrame, cols: List[str], filepath: Path,
                   masks: Dict[str, pd.Series]) -> Path:
        """
        Write formatted XLSX with material-type color coding.
        """
        # Default colors (same as entryops.py _export_single_file)
        steel_color = '#4a4a4a'
        aluminum_color = '#6495ED'
        copper_color = '#B87333'
        wood_color = '#8B4513'
        auto_color = '#2F4F4F'
        non232_color = '#FF0000'
        zero_metal_232_color = '#FF6600'  # Orange — cast iron / 9903.82.01
        font_color = '#000000'

        # Try to load user-specific colors from QSettings
        try:
            from PyQt5.QtCore import QSettings
            settings = QSettings("EntryOps", "EntryOps")
            steel_color = settings.value('export_steel_color', steel_color)
            aluminum_color = settings.value('export_aluminum_color', aluminum_color)
            copper_color = settings.value('export_copper_color', copper_color)
            wood_color = settings.value('export_wood_color', wood_color)
            auto_color = settings.value('export_automotive_color', auto_color)
            non232_color = settings.value('export_non232_color', non232_color)
            zero_metal_232_color = settings.value('export_zero_metal_232_color', zero_metal_232_color)
            font_color = settings.value('output_font_color', font_color)
        except Exception:
            pass

        font_color_rgb = '00' + font_color.lstrip('#').upper()
        steel_font = ExcelFont(name='Arial', size=11, color='00' + steel_color.lstrip('#').upper())
        aluminum_font = ExcelFont(name='Arial', size=11, color='00' + aluminum_color.lstrip('#').upper())
        copper_font = ExcelFont(name='Arial', size=11, color='00' + copper_color.lstrip('#').upper())
        wood_font = ExcelFont(name='Arial', size=11, color='00' + wood_color.lstrip('#').upper())
        auto_font = ExcelFont(name='Arial', size=11, color='00' + auto_color.lstrip('#').upper())
        non232_font = ExcelFont(name='Arial', size=11, color='00' + non232_color.lstrip('#').upper())
        zero_metal_232_font = ExcelFont(name='Arial', size=11, bold=True, color='00' + zero_metal_232_color.lstrip('#').upper())
        default_font = ExcelFont(name='Arial', size=11, color=font_color_rgb)
        center_alignment = Alignment(horizontal="center", vertical="center")

        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        purple_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")

        # Build index lists for each material type
        steel_indices = set(i for i, val in enumerate(masks['steel'].tolist()) if val)
        aluminum_indices = set(i for i, val in enumerate(masks['aluminum'].tolist()) if val)
        copper_indices = set(i for i, val in enumerate(masks['copper'].tolist()) if val)
        wood_indices = set(i for i, val in enumerate(masks['wood'].tolist()) if val)
        auto_indices = set(i for i, val in enumerate(masks['auto'].tolist()) if val)
        non232_indices = set(i for i, val in enumerate(masks['non232'].tolist()) if val)
        sec301_indices = set(i for i, val in enumerate(masks['sec301'].tolist()) if val)
        dual_dec_indices = set(i for i, val in enumerate(masks['dual_dec'].tolist()) if val)
        zero_metal_232_indices = set(i for i, val in enumerate(masks.get('zero_metal_232', pd.Series([False] * len(masks['steel']))).tolist()) if val)

        # Check if output is network path
        output_str = str(filepath)
        is_network = output_str.startswith('\\\\')

        if is_network:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                temp_path = Path(tmp.name)
            write_path = temp_path
        else:
            write_path = filepath

        with pd.ExcelWriter(write_path, engine='openpyxl') as writer:
            df_out[cols].to_excel(writer, index=False)
            ws = next(iter(writer.sheets.values()))
            ws.sheet_view.zoomScale = 85

            # Header row formatting
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = default_font
                cell.alignment = center_alignment

            # Data row formatting with material colors
            for row_num in range(2, len(df_out) + 2):
                row_idx = row_num - 2
                is_sec301 = row_idx in sec301_indices
                is_dual_dec = row_idx in dual_dec_indices

                # Cast iron / 9903.82.01 takes priority over the
                # steel/aluminum HTS classification color.
                if row_idx in zero_metal_232_indices:
                    font_to_use = zero_metal_232_font
                elif row_idx in steel_indices:
                    font_to_use = steel_font
                elif row_idx in aluminum_indices:
                    font_to_use = aluminum_font
                elif row_idx in copper_indices:
                    font_to_use = copper_font
                elif row_idx in wood_indices:
                    font_to_use = wood_font
                elif row_idx in auto_indices:
                    font_to_use = auto_font
                elif row_idx in non232_indices:
                    font_to_use = non232_font
                else:
                    font_to_use = default_font

                for col_idx in range(1, len(cols) + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = font_to_use
                    cell.alignment = center_alignment
                    if is_dual_dec:
                        cell.fill = purple_fill
                    elif is_sec301:
                        cell.fill = orange_fill

            # Auto-size columns
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = ws.cell(row=1, column=col_idx).column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            # Page setup
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        if is_network:
            import time as _time
            for attempt in range(3):
                try:
                    shutil.copy2(temp_path, filepath)
                    temp_path.unlink()
                    break
                except (OSError, PermissionError) as e:
                    if attempt < 2:
                        self._log(f"  Network write attempt {attempt + 1} failed, retrying in 1s...")
                        _time.sleep(1)
                    else:
                        self._log(f"  FAILED to copy to network path after 3 attempts: {e}")
                        self._log(f"  Temp file preserved at: {temp_path}")
                        filepath = temp_path  # Return temp path so user can recover

        self._log(f"  Exported: {filepath.name}")
        return filepath

    def export(self, df_enriched: pd.DataFrame, profile_name: str,
               filename: str, split_by_invoice: bool = None) -> List[Path]:
        """
        Full export pipeline: load profile -> build export df -> write XLSX.

        Args:
            df_enriched: Enriched DataFrame from EnrichmentPipeline
            profile_name: Name of output profile to use (or empty for default)
            filename: Base filename for output
            split_by_invoice: Override split setting (None = use profile setting)

        Returns:
            List of created XLSX file paths
        """
        # Load profile (or use defaults)
        profile = self.load_profile(profile_name) if profile_name else None
        if profile is None:
            profile = {
                'column_mapping': {name: name for name in DEFAULT_OUTPUT_COLUMN_ORDER},
                'column_order': DEFAULT_OUTPUT_COLUMN_ORDER.copy(),
                'column_visibility': {},
                'split_by_invoice': False,
            }

        if split_by_invoice is None:
            split_by_invoice = profile.get('split_by_invoice', False)

        # Add _invoice_number for split support
        if 'invoice_number' in df_enriched.columns:
            df_enriched['_invoice_number'] = df_enriched['invoice_number'].fillna('').astype(str)

        # Add _sec301_exclusion
        if 'Sec301_Exclusion_Tariff' in df_enriched.columns:
            df_enriched['_sec301_exclusion'] = df_enriched['Sec301_Exclusion_Tariff'].fillna('').astype(str)
        elif '_sec301_exclusion' not in df_enriched.columns:
            df_enriched['_sec301_exclusion'] = ''

        # Build export dataframe
        df_out, cols, masks = self.build_export_df(df_enriched, profile)

        created_files = []

        # Handle split by invoice
        if split_by_invoice and '_invoice_number' in df_out.columns:
            unique_invoices = df_out['_invoice_number'].dropna().unique()
            unique_invoices = [inv for inv in unique_invoices if inv and str(inv).strip() not in ['', 'nan', 'None']]

            if len(unique_invoices) > 1:
                for invoice_num in unique_invoices:
                    invoice_df = df_out[df_out['_invoice_number'] == invoice_num].copy()

                    # Recalculate masks for subset
                    inv_masks = {}
                    for key, mask in masks.items():
                        inv_mask_indices = invoice_df.index
                        inv_masks[key] = mask.reindex(inv_mask_indices).fillna(False).reset_index(drop=True)

                    invoice_df = invoice_df.reset_index(drop=True)
                    inv_filename = f"{invoice_num}_{datetime.now():%Y%m%d}.xlsx"
                    filepath = self.output_dir / inv_filename
                    self.write_xlsx(invoice_df, cols, filepath, inv_masks)
                    created_files.append(filepath)

                return created_files

        # Single file export
        filepath = self.output_dir / filename
        self.write_xlsx(df_out, cols, filepath, masks)
        created_files.append(filepath)

        return created_files
